from openpyxl import load_workbook
from openpyxl import Workbook

wb = load_workbook("userDB.xlsx")
ws = wb.active

def func():
    for row in range(2,100):
        if ws.cell(row,1).value is None:
            return row
            break
a = func()
print(a)

wb.save('userDB.xlsx')
