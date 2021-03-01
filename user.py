from openpyxl import load_workbook, Workbook

c_name = 1
c_id = 2
c_lvl = 3
c_exp = 4
c_money = 5
c_loss = 6

default_money = 10000

wb = load_workbook("userDB.xlsx")
ws = wb.active

def loadFile():
    wb = load_workbook("userDB.xlsx")
    ws = wb.active
def saveFile():
    wb.save("userDB.xlsx")
    wb.close()

#=========================checking==================================
def checkUserNum():
    print("user.py - checkUserNum")
    loadFile()

    count = 0

    for row in range(2, ws.max_row+1):
        if ws.cell(row,c_name).value != None:
            count = count+1
        else:
            count = count
    return count
def checkFirstRow():
    print("user.py - checkFirstRow")
    loadFile()

    print("첫번째 빈 곳을 탐색")

    for row in range(2, ws.max_row + 1):
        if ws.cell(row,1).value is None:
            return row
            break

    _result = ws.max_row+1

    saveFile()

    return _result

def checkUser(_name, _id):
    print("user.py - checkUser")
    print(str(_name) + "<" + str(_id) + ">의 존재 여부 확인")
    print("")

    loadFile()

    userNum = checkUserNum()
    print("등록된 유저수: ", userNum)
    print("")

    print("이름과 고유번호 탐색")
    print("")

    for row in range(2, 3+userNum):
        print(row, "번째 줄 name: ", ws.cell(row,c_name).value)
        print("입력된 name: ", _name)
        print("이름과 일치 여부: ", ws.cell(row, c_name).value == _name)

        print(row,"번째 줄 id: ", ws.cell(row,c_id).value)
        print("입력된 id: ", hex(_id))
        print("고유번호정보와 일치 여부: ", ws.cell(row, c_id).value == hex(_id))
        print("")

        if ws.cell(row, c_name).value == _name and ws.cell(row,c_id).value == hex(_id):
            print("등록된  이름과 고유번호를 발견")
            print("등록된  값의 위치: ",  row, "번째 줄")
            print("")

            saveFile()

            return True, row
            break
        else:
            print("등록된 정보를 탐색 실패, 재탐색 실시")

    saveFile()
    print("발견 실패")

    return False, None

#=========================Money==================================
def getMoney(_name,_row):
    print("user.py - getMoney")
    loadFile()

    print(_name, "의 돈을 탐색")

    result = ws.cell(_row, c_money).value
    print(_name,"의 보유 자산: ", result)

    saveFile()

    return result

def remit(sender, sender_row, receiver, receiver_row, _amount):
    print("user.py - remit")
    #loadFile()
    
    print("보내는 사람: ", sender)
    print("받는 사람: ", receiver)
    print("보내는 돈: ", _amount)
    print("")

    modifyMoney(receiver, receiver_row, int(_amount))
    modifyMoney(sender, sender_row, -int(_amount))

    print("")

def modifyMoney(_target, _row, _amount):
    print("user.py - modifyMoney")
    loadFile()

    print(_target, "의 자산데이터 수정")
    print(_target, "의 자산: " + str(ws.cell(_row, c_money).value))
    print("추가할 액수: ", _amount)
    ws.cell(_row, c_money).value += _amount

    print("자산데이터 수정 완료")
    print("수정된", _target, "의 자산: ", ws.cell(_row, c_money).value)
    
    saveFile()

def addLoss(_target, _row, _amount):
    print("user.py - addLoss")
    loadFile()

    print(_target, "의 잃은 돈 추가")
    print(_target, "의 잃은돈: " + str(ws.cell(_row, c_loss).value))
    print("추가할 액수: ", _amount)
    ws.cell(_row, c_loss).value += _amount

    print("잃은 돈 추가 완료")
    print(_target, "의 총 잃은 돈: ", ws.cell(_row, c_loss).value)

    saveFile()
#=========================Level==================================
def levelupCheck(_row):
    print("user.py - levelupCheck")
    loadFile()

    name = ws.cell(_row, c_name).value
    exp = ws.cell(_row, c_exp).value
    lvl = ws.cell(_row, c_lvl).value
    amount_to_up = lvl*lvl + 6*lvl
    count = 0

    print(name,"의 레벨업 조사")
    print(name, "의 현재 레벨: ", lvl, "(", exp, "/", amount_to_up,")")

    if exp >= amount_to_up:
        while(exp >= amount_to_up and exp >= 0):
            print("레벨업에 필요한 경험치 :", amount_to_up)
            print("현재 경험치: ", exp)

            print("충분한 경험치양을 확인")
            ws.cell(_row, c_lvl).value += 1
            count += 1
            print("레벨 데이터 수정")
            ws.cell(_row, c_exp).value -= amount_to_up
            print("경험치 초기화")

            lvl = ws.cell(_row, c_lvl).value
            exp = ws.cell(_row, c_exp).value
            amount_to_up = lvl*lvl + 6*lvl
        return True, lvl
    else:
        return False, lvl

def modifyExp(_row, _amount):
    print("user.py - modifyExp")
    loadFile()

    name = ws.cell(_row, c_name).value
    print(name, "의 경험치 획득량: ", _amount)

    ws.cell(_row, c_exp).value += _amount
    print(name, "현재 경험치: ", ws.cell(_row, c_exp).value)

    saveFile()
#=========================Ranking==================================
def ranking():
    print("user.py - ranking")

    loadFile()

    userRanking =  {}
    userNum = checkUserNum()

    print("등록된 유저수: ", userNum)
    print("")

    print("랭킹 집계중")

    for row in range(2, 2+userNum):
        _name = ws.cell(row, c_name).value
        _lvl = ws.cell(row, c_lvl).value
        userRanking[_name] = _lvl

    print("랭킹 집계 완료")
    a = sorted(userRanking.items(), reverse=True, key=lambda item:item[1])
    result = []
    for items in a:
        result.append(items[0])
        result.append(items[1])
    print(result)
    print("")

    return result

def getRank(_row):
    print("user.py - getRank")
    user = ws.cell(_row, c_name).value
    print(user, "의 랭킹 조사")
    rank = ranking()

    result = int(rank.index(user)/2)+1
    print(user, "의 랭킹: ",result, "위")

    return result

#=========================Account==================================
def Signup(_name, _id):
    print("user.py - signup")

    loadFile()

    _row = checkFirstRow()
    print("첫번째 빈곳: ", _row)
    print("")

    print("데이터 추가 시작")

    ws.cell(row=_row, column=c_name, value=_name)
    print("이름 추가 | ",  ws.cell(_row,c_name).value)
    ws.cell(row=_row, column=c_id, value =hex(_id))
    print("고유번호 추가 | ", ws.cell(_row,c_id).value)

    ws.cell(row=_row, column=c_lvl, value = 1)
    print("초기 레벨 설정 | lvl:", ws.cell(_row,c_lvl).value)
    ws.cell(row=_row, column=c_exp, value = 0)
    print("초기 경험치 설정 | exp:", ws.cell(_row,c_exp).value)

    ws.cell(row=_row, column=c_money, value = default_money)
    print("기본 자금 지급 | ", ws.cell(_row,c_money).value)
    ws.cell(row=_row, column=c_loss, value = 0)
    print("초기 손실 설정 | loss:", ws.cell(_row,c_loss).value)

    print("")

    saveFile()

    print("데이터 추가 완료")

def DeleteAccount(_row):
    print("user.py - DeleteAccount")
    loadFile()
    print("회원탈퇴 진행")

    print("유저 데이터 삭제")
    ws.delete_rows(_row)

    saveFile()
    
    print("회원탈퇴 완료")

def userInfo(_row):
    loadFile()

    _lvl = ws.cell(_row,c_lvl).value
    _exp = ws.cell(_row,c_exp).value
    _money = ws.cell(_row,c_money).value
    _loss = ws.cell(_row,c_loss).value

    print("레벨: ", _lvl)
    print("경험치: ", _exp)
    print("보유자산: ", _money)
    print("잃은 돈: ", _loss)

    saveFile()

    return _lvl, _exp, _money, _loss


#=========================For Test==================================
def resetData():
    loadFile()

    print("유저 데이터를 삭제")

    ws.delete_rows(2,ws.max_row)
    saveFile()

    print("데이터 삭제 완료")

def addMoney(_row, _amount):
    loadFile()

    ws.cell(_row, c_money).value += _amount

    saveFile()

def addExp(_row, _amount):
    loadFile()

    ws.cell(_row, c_exp).value += _amount

    saveFile()

def adjustlvl(_row, _amount):
    loadFile()
    
    ws.cell(_row, c_lvl).value = _amount
    ws.cell(_row, c_exp).value = 0

    saveFile
